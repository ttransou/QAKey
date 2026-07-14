"""Flask application entry-point for QAKey."""

from __future__ import annotations

import csv
import json
import io
import os
import uuid
from functools import wraps
from datetime import datetime, timezone
from typing import Optional

import yaml
from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook, load_workbook

from qakey.engine import QAEngine
from qakey.ingest import chunk_text, parse_chunks
from qakey.models import QARecord
from qakey.rendering import render_answer_html
from qakey.store import QAStore

# ---------------------------------------------------------------------------
# Initialise
# ---------------------------------------------------------------------------

app = Flask(__name__)
app.secret_key = os.environ.get("QAKEY_SECRET_KEY", "qakey-dev-secret")

# Prevent stale UI during iterative editing by reloading templates/static assets.
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0
app.jinja_env.auto_reload = True

_CONFIG_PATH = os.environ.get("QAKEY_CONFIG", "config.yaml")

with open(_CONFIG_PATH, encoding="utf-8") as _fh:
    _config: dict = yaml.safe_load(_fh)

_store = QAStore(_config["knowledge"]["records_path"])
_matching_cfg = _config.get("matching", {})

_synonyms: dict = {}
_synonyms_path = _config["knowledge"].get("synonyms_path", "")
if _synonyms_path and os.path.exists(_synonyms_path):
    with open(_synonyms_path, encoding="utf-8") as _fh:
        _syn_data = yaml.safe_load(_fh) or {}
    _synonyms = _syn_data.get("synonyms", {})

_engine = QAEngine(
    records=_store.get_active(),
    synonyms=_synonyms,
    confidence_threshold=_matching_cfg.get("confidence_threshold", 0.25),
    ambiguity_margin=_matching_cfg.get("ambiguity_margin", 0.08),
    max_suggestions=_matching_cfg.get("max_suggestions", 3),
    no_match_message=_config.get("fallback", {}).get(
        "no_match_message",
        "I could not find an approved answer for that question. Please rephrase your question or contact the appropriate team directly.",
    ),
    ambiguous_match_message=_config.get("fallback", {}).get(
        "ambiguous_match_message",
        "I found more than one possible approved question. Please choose the closest match or rephrase your question.",
    ),
)

_editor_cfg = _config.get("editor", {})
_fallback_cfg = _config.get("fallback", {})
_valid_statuses = {"Draft", "Active", "Inactive"}
_IMPORT_TEMPLATE_HEADERS = [
    "canonical_question",
    "answer",
    "status",
    "alternate_phrasings",
    "tags",
    "contributor",
    "reviewer",
]


def _editor_auth_enabled() -> bool:
    return bool(_editor_cfg.get("require_auth", False))


def _editor_username() -> str:
    return os.environ.get("QAKEY_EDITOR_USERNAME", _editor_cfg.get("admin_username", "admin"))


def _editor_password() -> str:
    return os.environ.get("QAKEY_EDITOR_PASSWORD", _editor_cfg.get("admin_password", ""))


def _is_editor_authenticated() -> bool:
    if not _editor_auth_enabled():
        return True
    return bool(session.get("editor_authenticated", False))


def _require_editor_page_auth(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if _is_editor_authenticated():
            return func(*args, **kwargs)
        return redirect(url_for("editor_login", next=request.path))

    return wrapper


def _require_editor_api_auth(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if _is_editor_authenticated():
            return func(*args, **kwargs)
        return jsonify({"error": "editor authentication required"}), 401

    return wrapper


def _split_multi_value(text: str) -> list[str]:
    if not text:
        return []
    raw = text.replace("\r", "\n")
    for sep in ("|", ";"):
        raw = raw.replace(sep, "\n")
    return [part.strip() for part in raw.split("\n") if part.strip()]


def _normalize_field_name(name: str) -> str:
    return str(name).strip().lower().replace(" ", "_").replace("-", "_")


def _parse_records_from_csv(raw: bytes, defaults: dict) -> tuple[list[dict], list[str]]:
    rows: list[dict] = []
    errors: list[str] = []

    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return [], ["Missing header row in CSV file"]

    field_map = {_normalize_field_name(name): name for name in reader.fieldnames}

    def get_value(row: dict, *keys: str) -> str:
        for key in keys:
            source = field_map.get(_normalize_field_name(key))
            if source and row.get(source) is not None:
                return str(row.get(source)).strip()
        return ""

    for row_index, row in enumerate(reader, start=2):
        canonical = get_value(row, "canonical_question", "question")
        answer = get_value(row, "answer")
        if not canonical or not answer:
            errors.append(f"Row {row_index}: canonical_question and answer are required")
            continue

        status = get_value(row, "status") or defaults["status"]
        if status not in _valid_statuses:
            status = defaults["status"]

        rows.append(
            {
                "canonical_question": canonical,
                "answer": answer,
                "status": status,
                "alternate_phrasings": _split_multi_value(get_value(row, "alternate_phrasings", "alternates")),
                "tags": _split_multi_value(get_value(row, "tags")) or defaults["tags"],
                "contributor": get_value(row, "contributor") or defaults["contributor"],
                "reviewer": get_value(row, "reviewer") or defaults["reviewer"],
            }
        )

    return rows, errors


def _parse_records_from_xlsx(raw: bytes, defaults: dict) -> tuple[list[dict], list[str]]:
    rows: list[dict] = []
    errors: list[str] = []

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    data_rows = list(sheet.iter_rows(values_only=True))
    if not data_rows:
        return [], ["Workbook is empty"]

    headers = [str(cell).strip() if cell is not None else "" for cell in data_rows[0]]
    header_map = {_normalize_field_name(header): idx for idx, header in enumerate(headers) if header}

    def cell_value(row: tuple, *keys: str) -> str:
        for key in keys:
            idx = header_map.get(_normalize_field_name(key))
            if idx is not None and idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
        return ""

    for row_number, row in enumerate(data_rows[1:], start=2):
        canonical = cell_value(row, "canonical_question", "question")
        answer = cell_value(row, "answer")
        if not canonical and not answer:
            continue
        if not canonical or not answer:
            errors.append(f"Row {row_number}: canonical_question and answer are required")
            continue

        status = cell_value(row, "status") or defaults["status"]
        if status not in _valid_statuses:
            status = defaults["status"]

        rows.append(
            {
                "canonical_question": canonical,
                "answer": answer,
                "status": status,
                "alternate_phrasings": _split_multi_value(cell_value(row, "alternate_phrasings", "alternates")),
                "tags": _split_multi_value(cell_value(row, "tags")) or defaults["tags"],
                "contributor": cell_value(row, "contributor") or defaults["contributor"],
                "reviewer": cell_value(row, "reviewer") or defaults["reviewer"],
            }
        )

    return rows, errors


def _parse_import_records(file_name: str, raw: bytes, defaults: dict) -> tuple[list[dict], list[str]]:
    lower_name = file_name.lower()
    if lower_name.endswith(".csv"):
        return _parse_records_from_csv(raw, defaults)
    if lower_name.endswith(".xlsx"):
        return _parse_records_from_xlsx(raw, defaults)
    return [], ["Unsupported file type. Only .csv and .xlsx are allowed"]


def _build_import_template_csv() -> io.BytesIO:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=_IMPORT_TEMPLATE_HEADERS)
    writer.writeheader()
    return io.BytesIO(output.getvalue().encode("utf-8"))


def _build_import_template_xlsx() -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "QAKey Import Template"
    sheet.append(_IMPORT_TEMPLATE_HEADERS)
    sheet.freeze_panes = "A2"
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _feedback_log_path() -> str:
    return _fallback_cfg.get("fallback_log_path", "")


def _feedback_logging_enabled() -> bool:
    return bool(_fallback_cfg.get("enabled", True)) and bool(_feedback_log_path())


def _normalize_contact_route(route_cfg: object, default_label: str) -> Optional[dict]:
    if not isinstance(route_cfg, dict):
        return None

    if not bool(route_cfg.get("enabled", True)):
        return None

    value = (
        route_cfg.get("value")
        or route_cfg.get("email")
        or route_cfg.get("url")
        or ""
    ).strip()
    if not value:
        return None

    route_type = str(route_cfg.get("type") or "").strip().lower()
    if not route_type:
        route_type = "url" if value.startswith(("http://", "https://")) else "email"

    label = str(route_cfg.get("label") or default_label).strip() or default_label
    display_text = str(route_cfg.get("display_text") or "").strip() or label
    href = value if route_type == "url" else value.removeprefix("mailto:")
    if route_type == "email" and not href.startswith("mailto:"):
        href = f"mailto:{href}"

    return {
        "enabled": True,
        "label": label,
        "display_text": display_text,
        "type": route_type,
        "value": value,
        "href": href,
    }


def _fallback_human_help() -> Optional[dict]:
    route = _normalize_contact_route(_fallback_cfg.get("human_help"), "Contact the team")
    if route is not None:
        return route

    fallback_routes = _fallback_cfg.get("fallback_routes", {})
    if isinstance(fallback_routes, dict):
        return _normalize_contact_route(fallback_routes.get("default"), "Contact the team")

    return None


def _load_feedback_alerts() -> list[dict]:
    log_path = _feedback_log_path()
    if not log_path:
        return []

    if not os.path.exists(log_path):
        return []

    with open(log_path, encoding="utf-8") as fh:
        raw = fh.read().strip()
    if not raw:
        return []

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        alerts: list[dict] = []
        for line in raw.splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                alerts.append(json.loads(line))
            except json.JSONDecodeError:
                continue
        return alerts

    if isinstance(data, dict):
        alerts = data.get("alerts") or []
        if isinstance(alerts, list):
            return alerts

    if isinstance(data, list):
        return data

    return []


def _save_feedback_alerts(alerts: list[dict]) -> None:
    log_path = _feedback_log_path()
    if not log_path:
        return

    os.makedirs(os.path.dirname(log_path) or ".", exist_ok=True)
    compact_alerts = alerts[-int(_fallback_cfg.get("max_alerts", 25)):]
    with open(log_path, "w", encoding="utf-8") as fh:
        json.dump({"alerts": compact_alerts}, fh, ensure_ascii=False, separators=(",", ":"))


def _upsert_feedback_alert(event: dict) -> bool:
    alerts = _load_feedback_alerts()
    key = (
        event.get("question", ""),
        event.get("record_id") or "",
        event.get("fallback_type") or "",
    )

    for alert in alerts:
        alert_key = (
            alert.get("question", ""),
            alert.get("record_id") or "",
            alert.get("fallback_type") or "",
        )
        if alert_key == key and alert.get("resolved_at") is None:
            alert["occurrences"] = int(alert.get("occurrences", 1)) + 1
            alert["last_seen_at"] = event["timestamp"]
            alert["helpful"] = False
            _save_feedback_alerts(alerts)
            return True

    alert = {
        "id": str(uuid.uuid4()),
        "created_at": event["timestamp"],
        "last_seen_at": event["timestamp"],
        "question": event.get("question", ""),
        "matched": bool(event.get("matched", False)),
        "record_id": event.get("record_id") or None,
        "fallback_type": event.get("fallback_type") or None,
        "confidence": event.get("confidence"),
        "occurrences": 1,
        "helpful": False,
        "comment": event.get("comment") or "",
        "resolved_at": None,
    }
    alerts.append(alert)
    _save_feedback_alerts(alerts)
    return True


def _resolve_feedback_alert(alert_id: str) -> bool:
    alerts = _load_feedback_alerts()
    changed = False
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    for alert in alerts:
        if alert.get("id") == alert_id and alert.get("resolved_at") is None:
            alert["resolved_at"] = now
            changed = True
            break

    if changed:
        alerts = [alert for alert in alerts if alert.get("resolved_at") is None]
        _save_feedback_alerts(alerts)
    return changed

# ---------------------------------------------------------------------------
# Page routes
# ---------------------------------------------------------------------------


@app.route("/")
def index():
    return render_template("index.html", config=_config)


@app.route("/editor")
@_require_editor_page_auth
def editor():
    return render_template("editor.html", config=_config)


@app.route("/editor/login", methods=["GET", "POST"])
def editor_login():
    if not _editor_auth_enabled():
        return redirect(url_for("editor"))

    if _is_editor_authenticated():
        return redirect(url_for("editor"))

    error = ""
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""

        if username == _editor_username() and password == _editor_password() and password:
            session["editor_authenticated"] = True
            next_url = request.args.get("next") or url_for("editor")
            return redirect(next_url)
        error = "Invalid admin credentials"

    return render_template("editor_login.html", config=_config, error=error)


@app.route("/editor/logout", methods=["POST"])
def editor_logout():
    session.pop("editor_authenticated", None)
    if _editor_auth_enabled():
        return redirect(url_for("editor_login"))
    return redirect(url_for("editor"))


# ---------------------------------------------------------------------------
# API routes
# ---------------------------------------------------------------------------


@app.route("/api/query", methods=["POST"])
def api_query():
    """Match a user question and return the canonical answer."""
    body = request.get_json(silent=True) or {}
    question = (body.get("question") or "").strip()
    if not question:
        return jsonify({"error": "question is required"}), 400

    result = _engine.match(question)
    payload = result.to_dict()
    if payload.get("matched") and payload.get("answer"):
        payload["answer_html"] = render_answer_html(payload["answer"])
    elif payload.get("answer"):
        payload["human_help"] = _fallback_human_help()
    return jsonify(payload)


@app.route("/api/feedback", methods=["POST"])
def api_feedback():
    """Record simple thumbs-up/down feedback for a query result."""
    body = request.get_json(silent=True) or {}
    question = (body.get("question") or "").strip()
    if not question:
        return jsonify({"error": "question is required"}), 400

    helpful = body.get("helpful")
    if not isinstance(helpful, bool):
        return jsonify({"error": "helpful must be a boolean"}), 400

    event = {
        "timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "question": question,
        "helpful": helpful,
        "matched": bool(body.get("matched", False)),
        "record_id": (body.get("record_id") or None),
        "fallback_type": (body.get("fallback_type") or None),
        "confidence": body.get("confidence"),
        "comment": (body.get("comment") or "").strip(),
    }

    recorded = False
    if _feedback_logging_enabled() and (not helpful or not event["matched"]):
        recorded = _upsert_feedback_alert(event)

    return jsonify({"success": True, "recorded": recorded})


@app.route("/api/editor/feedback-alerts", methods=["GET"])
@_require_editor_api_auth
def api_editor_feedback_alerts():
    alerts = _load_feedback_alerts()
    return jsonify({"count": len(alerts), "alerts": alerts})


@app.route("/api/editor/feedback-alerts/<alert_id>/resolve", methods=["POST"])
@_require_editor_api_auth
def api_editor_feedback_alert_resolve(alert_id: str):
    if not _resolve_feedback_alert(alert_id):
        return jsonify({"error": "alert not found"}), 404
    return jsonify({"success": True})


@app.route("/api/records", methods=["GET"])
@_require_editor_api_auth
def api_list_records():
    return jsonify([r.to_dict() for r in _store.get_all()])


@app.route("/api/records", methods=["POST"])
@_require_editor_api_auth
def api_create_record():
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "request body required"}), 400

    record = QARecord.from_dict(body)
    _store.create(record)
    return jsonify(record.to_dict()), 201


@app.route("/api/records/<record_id>", methods=["PUT"])
@_require_editor_api_auth
def api_update_record(record_id: str):
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "request body required"}), 400

    existing = _store.get(record_id)
    if existing is None:
        return jsonify({"error": "record not found"}), 404

    body["id"] = record_id
    body["created_at"] = existing.created_at
    record = QARecord.from_dict(body)
    _store.update(record)
    return jsonify(record.to_dict())


@app.route("/api/records/<record_id>", methods=["DELETE"])
@_require_editor_api_auth
def api_delete_record(record_id: str):
    if not _store.delete(record_id):
        return jsonify({"error": "record not found"}), 404
    return "", 204


@app.route("/api/records/import-preview", methods=["POST"])
@_require_editor_api_auth
def api_import_preview():
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return jsonify({"error": "file is required"}), 400

    file_name = upload.filename
    if not file_name.lower().endswith((".csv", ".xlsx")):
        return jsonify({"error": "Only .csv and .xlsx files are supported"}), 400

    defaults = {
        "status": (request.form.get("default_status") or "Draft").strip(),
        "contributor": (request.form.get("default_contributor") or "").strip(),
        "reviewer": (request.form.get("default_reviewer") or "").strip(),
        "tags": _split_multi_value(request.form.get("default_tags") or ""),
    }
    if defaults["status"] not in _valid_statuses:
        defaults["status"] = "Draft"

    raw = upload.read()
    records, errors = _parse_import_records(file_name, raw, defaults)

    return jsonify(
        {
            "count": len(records),
            "errors": errors,
            "records": records,
        }
    )


@app.route("/api/records/export", methods=["GET"])
@_require_editor_api_auth
def api_records_export():
    export_format = (request.args.get("format") or "csv").lower()
    records = _store.get_all()

    headers = [
        "id",
        "canonical_question",
        "answer",
        "status",
        "alternate_phrasings",
        "tags",
        "contributor",
        "reviewer",
        "created_at",
        "updated_at",
        "version",
    ]

    if export_format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        for record in records:
            payload = record.to_dict()
            payload["alternate_phrasings"] = "\n".join(payload.get("alternate_phrasings") or [])
            payload["tags"] = "\n".join(payload.get("tags") or [])
            writer.writerow({key: payload.get(key, "") for key in headers})

        csv_bytes = io.BytesIO(output.getvalue().encode("utf-8"))
        return send_file(
            csv_bytes,
            mimetype="text/csv",
            as_attachment=True,
            download_name="qakey-records.csv",
        )

    if export_format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "QAKey Records"
        sheet.append(headers)
        for record in records:
            payload = record.to_dict()
            payload["alternate_phrasings"] = "\n".join(payload.get("alternate_phrasings") or [])
            payload["tags"] = "\n".join(payload.get("tags") or [])
            sheet.append([payload.get(key, "") for key in headers])

        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return send_file(
            stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="qakey-records.xlsx",
        )

    return jsonify({"error": "format must be 'csv' or 'xlsx'"}), 400


@app.route("/api/records/import-template", methods=["GET"])
@_require_editor_api_auth
def api_records_import_template():
    template_format = (request.args.get("format") or "csv").lower()

    if template_format == "csv":
        csv_bytes = _build_import_template_csv()
        return send_file(
            csv_bytes,
            mimetype="text/csv",
            as_attachment=True,
            download_name="qakey-import-template.csv",
        )

    if template_format == "xlsx":
        stream = _build_import_template_xlsx()
        return send_file(
            stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="qakey-import-template.xlsx",
        )

    return jsonify({"error": "format must be 'csv' or 'xlsx'"}), 400


@app.route("/api/ingest/preview", methods=["POST"])
@_require_editor_api_auth
def api_ingest_preview():
    """Deterministically extract, chunk, and parse raw source text."""
    body = request.get_json(silent=True) or {}
    text = (body.get("text") or "").strip()
    if not text:
        return jsonify({"error": "text is required"}), 400

    status = body.get("status") or "Draft"
    contributor = (body.get("contributor") or "").strip()
    reviewer = (body.get("reviewer") or "").strip()
    tags = [str(tag).strip() for tag in body.get("tags") or [] if str(tag).strip()]
    max_chars = int(body.get("max_chars") or 650)

    chunks = chunk_text(text, max_chars=max_chars)
    candidates = parse_chunks(chunks)
    records = [
        candidate.to_record(
            status=status,
            contributor=contributor,
            reviewer=reviewer,
            tags=tags,
        )
        for candidate in candidates
    ]

    preview_records = []
    for record in records:
        record_payload = record.to_dict()
        record_payload["answer_html"] = render_answer_html(record.answer)
        preview_records.append(record_payload)

    return jsonify(
        {
            "count": len(records),
            "chunks": [
                {"index": idx, "text": chunk}
                for idx, chunk in enumerate(chunks)
            ],
            "records": preview_records,
            "candidates": [candidate.to_dict() for candidate in candidates],
        }
    )


@app.route("/api/publish", methods=["POST"])
@_require_editor_api_auth
def api_publish():
    """Validate all records, persist to YAML, and rebuild the index."""
    result = _store.publish()
    if result["success"]:
        _engine.rebuild(_store.get_active())
    return jsonify(result)


# ---------------------------------------------------------------------------
# Health-check
# ---------------------------------------------------------------------------


@app.route("/api/health")
def api_health():
    return jsonify({"status": "ok", "version": "1.0.0"})


# ---------------------------------------------------------------------------
# Entry-point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app.run(
        debug=_config.get("debug", False),
        host=_config.get("host", "127.0.0.1"),
        port=_config.get("port", 5000),
    )
