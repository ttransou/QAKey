"""Tests for downloadable import templates."""

from io import BytesIO

from openpyxl import load_workbook

from app import app


EXPECTED_HEADERS = [
    "canonical_question",
    "answer",
    "status",
    "alternate_phrasings",
    "tags",
    "contributor",
    "reviewer",
]


def test_import_template_csv_matches_import_schema():
    client = app.test_client()

    response = client.get("/api/records/import-template?format=csv")

    assert response.status_code == 200
    assert response.mimetype == "text/csv"
    assert "qakey-import-template.csv" in response.headers["Content-Disposition"]

    lines = response.data.decode("utf-8").splitlines()
    assert lines[0].split(",") == EXPECTED_HEADERS


def test_import_template_xlsx_matches_import_schema():
    client = app.test_client()

    response = client.get("/api/records/import-template?format=xlsx")

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "qakey-import-template.xlsx" in response.headers["Content-Disposition"]

    workbook = load_workbook(BytesIO(response.data), read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert headers == EXPECTED_HEADERS
